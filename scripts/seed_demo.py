#!/usr/bin/env python3
"""
CloudCampus — Full Demo Seed Script
Provisions 4 school tenants with bulk student/teacher/class data + user accounts.
"""

import io
import json
import sys
import time
import warnings

warnings.filterwarnings("ignore")

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "http://localhost:8080/api/v1"

# ── Credentials ────────────────────────────────────────────────────────────────
SA_USER = "superadmin"
SA_PASS = "SuperAdmin_Docker_2026!"

# ── Demo Tenants ───────────────────────────────────────────────────────────────
TENANTS = [
    {
        "tenantId":    "sunrise-academy",
        "schoolName":  "Sunrise Academy",
        "schemaName":  "school_sunrise-academy",
        "logoUrl":     "https://images.unsplash.com/photo-1513258496099-48168024aec0?auto=format&fit=crop&w=200&q=80",
        "primaryColor": "#10b981",
        "admin":       {"fullName": "Priya Sharma",    "username": "priya.sharma",  "email": "priya@sunrise.edu",   "password": "Sunrise@2026!"},
        "code":        "SUN",
    },
    {
        "tenantId":    "greenwood-high",
        "schoolName":  "Greenwood High",
        "schemaName":  "school_greenwood-high",
        "logoUrl":     "https://images.unsplash.com/photo-1503676260728-1c00da094a0b?auto=format&fit=crop&w=200&q=80",
        "primaryColor": "#2563eb",
        "admin":       {"fullName": "Arjun Mehta",     "username": "arjun.mehta",   "email": "arjun@greenwood.edu", "password": "Greenwood@2026!"},
        "code":        "GWH",
    },
    {
        "tenantId":    "riverdale-public",
        "schoolName":  "Riverdale Public School",
        "schemaName":  "school_riverdale-public",
        "logoUrl":     "https://images.unsplash.com/photo-1580582932707-520aed937b7b?auto=format&fit=crop&w=200&q=80",
        "primaryColor": "#7c3aed",
        "admin":       {"fullName": "Kavya Nair",      "username": "kavya.nair",    "email": "kavya@riverdale.edu", "password": "Riverdale@2026!"},
        "code":        "RVD",
    },
    {
        "tenantId":    "oakridge-international",
        "schoolName":  "Oakridge International",
        "schemaName":  "school_oakridge-international",
        "logoUrl":     "https://images.unsplash.com/photo-1523050854058-8df90110c9f1?auto=format&fit=crop&w=200&q=80",
        "primaryColor": "#f59e0b",
        "admin":       {"fullName": "Rohan Kapoor",    "username": "rohan.kapoor",  "email": "rohan@oakridge.edu",  "password": "Oakridge@2026!"},
        "code":        "OAK",
    },
]

# ── Per-school template data ───────────────────────────────────────────────────
# classes: (name, code_suffix)
CLASSES = [
    ("Grade 1",  "G1"),
    ("Grade 3",  "G3"),
    ("Grade 5",  "G5"),
    ("Grade 7",  "G7"),
    ("Grade 10", "G10"),
]
# sections per class
SECTIONS = ["A", "B"]

# 15 students per school — names cycled
STUDENT_FIRST = ["Aarav","Diya","Kabir","Ananya","Vivaan","Ishita","Aryan","Priya",
                 "Rohan","Sneha","Aditya","Meera","Siddharth","Pooja","Karan"]
STUDENT_LAST  = ["Sharma","Patel","Singh","Kumar","Verma","Reddy","Nair","Joshi",
                 "Mehta","Gupta","Iyer","Rao","Shah","Chowdhury","Malhotra"]

# 6 teachers per school
TEACHER_DATA = [
    ("Sunita",  "Aggarwal",  "Math"),
    ("Vikram",  "Desai",     "Science"),
    ("Lakshmi", "Krishnan",  "English"),
    ("Rajesh",  "Bose",      "History"),
    ("Anita",   "Pillai",    "Geography"),
    ("Deepak",  "Chauhan",   "Physics"),
]

# 3 parents per school — linked to first 3 students
PARENT_DATA = [
    ("Ramesh",  "Sharma"),
    ("Sujata",  "Patel"),
    ("Vijay",   "Singh"),
]

# ── Helper ─────────────────────────────────────────────────────────────────────
def api(method, path, token=None, tenant=None, json_body=None, files=None):
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if tenant:
        headers["X-Tenant-ID"] = tenant
    if json_body is not None and files is None:
        headers["Content-Type"] = "application/json"

    resp = requests.request(
        method, f"{BASE}{path}",
        headers=headers,
        json=json_body,
        files=files,
        timeout=30,
    )
    try:
        return resp.json()
    except Exception:
        return {"success": False, "message": resp.text}

def ok(r):
    return r.get("success") or r.get("data") is not None

def login(username, password, tenant=None):
    payload = {"username": username, "password": password}
    if tenant:
        payload["tenantId"] = tenant
    headers = {"Content-Type": "application/json"}
    if tenant:
        headers["X-Tenant-ID"] = tenant
    r = requests.post(f"{BASE}/auth/login", json=payload, headers=headers, timeout=15)
    data = r.json()
    if data.get("success"):
        return data["data"]["accessToken"], data["data"].get("userId")
    return None, None

def pad(s, w=35):
    return str(s).ljust(w)

def log(symbol, msg):
    print(f"  {symbol} {msg}")

# ── Excel builder ──────────────────────────────────────────────────────────────
def build_workbook(code, students, teachers, classes, sections):
    wb = openpyxl.Workbook()

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    center       = Alignment(horizontal="center")

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = center
        for row in rows:
            ws.append(row)
        for col_idx, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        return ws

    add_sheet("STUDENTS", ["admission_no","first_name","last_name","dob","gender","email","phone"],
              students)
    add_sheet("TEACHERS", ["employee_no","first_name","last_name","email","phone","hire_date"],
              teachers)
    add_sheet("CLASSES",  ["class_name","class_code"],   classes)
    add_sheet("SECTIONS", ["section_name","class_code"], sections)

    # remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── Main seed ──────────────────────────────────────────────────────────────────
def seed():
    print("\n" + "═"*60)
    print("  CloudCampus — Demo Seed Script")
    print("═"*60)

    # ── 1. Super-admin login ───────────────────────────────────────
    print("\n[1/6] Authenticating as Super Admin …")
    sa_token, _ = login(SA_USER, SA_PASS)
    if not sa_token:
        print("  ❌  Could not log in as superadmin. Is the backend running?")
        sys.exit(1)
    print("  ✅  Logged in")

    # ── 2. Resolve FREE plan id ────────────────────────────────────
    plans_resp = api("GET", "/plans", token=sa_token)
    plans = plans_resp.get("data", [])
    basic_plan = next((p for p in plans if p["name"] == "BASIC"), None)
    free_plan  = next((p for p in plans if p["name"] == "FREE"),  None)
    plan       = basic_plan or free_plan
    if not plan:
        print("  ❌  No plan found. Aborting.")
        sys.exit(1)
    plan_id   = plan["id"]
    plan_name = plan["name"]
    print(f"\n[2/6] Subscription plan: {plan_name} (id={plan_id})")

    # ── 3. Per-tenant setup ────────────────────────────────────────
    all_credentials = []

    for idx, tenant in enumerate(TENANTS, 1):
        tid     = tenant["tenantId"]
        code    = tenant["code"]
        schema  = None

        print(f"\n{'─'*60}")
        print(f"[3/6] Tenant {idx}/4 — {tenant['schoolName']} ({tid})")
        print(f"{'─'*60}")

        # 3a. Create tenant (skip if exists)
        t_resp = api("POST", "/tenants", token=sa_token, json_body={
            "tenantId":    tid,
            "schoolName":  tenant["schoolName"],
            "primaryColor": tenant["primaryColor"],
            "logoUrl":     tenant["logoUrl"],
        })
        if ok(t_resp):
            schema = t_resp["data"]["schemaName"]
            log("✅", f"Tenant created  → schema={schema}")
        elif "already exists" in str(t_resp.get("message", "")).lower() or \
             "duplicate" in str(t_resp.get("message", "")).lower():
            # fetch existing schema
            g = api("GET", f"/tenants/{tid}", token=sa_token)
            if ok(g):
                schema = g["data"]["schemaName"]
                log("ℹ️ ", f"Tenant exists   → schema={schema}")
            else:
                log("❌", f"Cannot resolve schema for {tid}: {t_resp}")
                continue
        else:
            log("❌", f"Tenant creation failed: {t_resp}")
            continue

        # 3b. Subscribe
        sub_resp = api("POST", f"/tenants/{tid}/subscribe", token=sa_token,
                       json_body={"planId": plan_id, "durationDays": 365})
        if ok(sub_resp):
            log("✅", f"Subscribed to {plan_name}")
        else:
            msg = sub_resp.get("message", "")
            if "already" in msg.lower() or "active" in msg.lower():
                log("ℹ️ ", "Already subscribed")
            else:
                log("⚠️ ", f"Subscribe: {msg}")

        # 3c. Create school admin
        adm = tenant["admin"]
        u_resp = api("POST", "/users", token=sa_token, tenant=schema, json_body={
            "fullName": adm["fullName"],
            "username": adm["username"],
            "email":    adm["email"],
            "password": adm["password"],
            "role":     "SCHOOL_ADMIN",
        })
        if ok(u_resp):
            adm_user_id = u_resp["data"]["id"]
            log("✅", f"School admin created  ({adm['username']})")
        elif "already" in str(u_resp.get("message", "")).lower() or \
             "duplicate" in str(u_resp.get("message", "")).lower():
            log("ℹ️ ", f"Admin user exists ({adm['username']})")
            adm_user_id = None
        else:
            log("❌", f"Admin user: {u_resp}")
            adm_user_id = None

        # 3d. Login as school admin
        adm_token, adm_uid = login(adm["username"], adm["password"], tenant=schema)
        if not adm_token:
            log("❌", "Admin login failed — skipping bulk upload for this tenant")
            continue
        if adm_uid:
            adm_user_id = adm_uid
        log("✅", f"School admin logged in (userId={adm_user_id})")

        # ── Build Excel data ───────────────────────────────────────
        students_rows = []
        teachers_rows = []
        classes_rows  = []
        sections_rows = []
        student_creds = []
        teacher_creds = []

        for c_name, c_sfx in CLASSES:
            c_code = f"{code}-{c_sfx}"
            classes_rows.append([c_name, c_code])
            for sec in SECTIONS:
                sections_rows.append([sec, c_code])

        for i, (fn, ln) in enumerate(zip(STUDENT_FIRST, STUDENT_LAST), 1):
            adm_no  = f"{code}-S{i:03d}"
            email   = f"{fn.lower()}.{ln.lower()}{i}@{tid}.edu"
            phone   = f"98765{i:05d}"
            dob     = f"{2010 - (i % 5)}-{(i % 12)+1:02d}-{(i % 28)+1:02d}"
            gender  = "MALE" if i % 2 else "FEMALE"
            username = f"{fn.lower()}.{ln.lower()}{i}"
            password = f"{fn}@Student{i}26!"
            students_rows.append([adm_no, fn, ln, dob, gender, email, phone])
            student_creds.append({
                "name":     f"{fn} {ln}",
                "admNo":    adm_no,
                "username": username,
                "password": password,
                "email":    email,
                "role":     "STUDENT",
            })

        for i, (fn, ln, subject) in enumerate(TEACHER_DATA, 1):
            emp_no  = f"{code}-T{i:02d}"
            email   = f"{fn.lower()}.{ln.lower()}@{tid}.edu"
            phone   = f"91234{i:05d}"
            hire    = f"202{i % 4}-0{i}-15"
            username = f"{fn.lower()}.{ln.lower()}"
            password = f"{fn}@Teacher{i}26!"
            teachers_rows.append([emp_no, fn, ln, email, phone, hire])
            teacher_creds.append({
                "name":     f"{fn} {ln}",
                "empNo":    emp_no,
                "subject":  subject,
                "username": username,
                "password": password,
                "email":    email,
                "role":     "TEACHER",
            })

        # 3e. Bulk upload Excel
        xlsx_bytes = build_workbook(code, students_rows, teachers_rows, classes_rows, sections_rows)
        up_resp = api("POST", "/bulk/upload", token=adm_token, tenant=schema,
                      files={"file": ("bulk.xlsx", xlsx_bytes,
                                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        if ok(up_resp):
            d = up_resp.get("data", {})
            log("✅", f"Bulk upload → {d.get('successCount')}/{d.get('totalRows')} rows OK, "
                      f"{d.get('failedCount')} failed")
            if d.get("errors"):
                for err in d["errors"][:3]:
                    log("  ⚠️", f"  Row {err.get('rowNumber')} [{err.get('sheet')}]: {err.get('message')}")
        else:
            log("❌", f"Bulk upload failed: {up_resp.get('message')}")

        # 3f. Create teacher user accounts
        log("", "Creating teacher user accounts …")
        for tc in teacher_creds:
            u = api("POST", "/users", token=adm_token, tenant=schema, json_body={
                "fullName": tc["name"],
                "username": tc["username"],
                "email":    tc["email"],
                "password": tc["password"],
                "role":     "TEACHER",
            })
            if ok(u):
                tc["userId"] = u["data"]["id"]
                log("  ✅", f"{tc['username']}")
            elif "already" in str(u.get("message","")).lower() or \
                 "duplicate" in str(u.get("message","")).lower():
                log("  ℹ️", f"{tc['username']} (exists)")
            else:
                log("  ⚠️", f"{tc['username']}: {u.get('message')}")

        # 3g. Create student user accounts (first 5 only to keep it manageable)
        log("", "Creating student user accounts …")
        for sc in student_creds[:5]:
            u = api("POST", "/users", token=adm_token, tenant=schema, json_body={
                "fullName": sc["name"],
                "username": sc["username"],
                "email":    sc["email"],
                "password": sc["password"],
                "role":     "STUDENT",
            })
            if ok(u):
                sc["userId"] = u["data"]["id"]
                log("  ✅", f"{sc['username']}")
            elif "already" in str(u.get("message","")).lower() or \
                 "duplicate" in str(u.get("message","")).lower():
                log("  ℹ️", f"{sc['username']} (exists)")
            else:
                log("  ⚠️", f"{sc['username']}: {u.get('message')}")

        # 3h. Create parent user accounts and link to first 3 students
        log("", "Creating parent user accounts …")
        # Fetch student entity UUIDs (needed for parent–student link)
        stu_resp = api("GET", "/students?size=20&page=0", token=adm_token, tenant=schema)
        stu_entities = {}
        if ok(stu_resp):
            for s in stu_resp.get("data", {}).get("content", []):
                stu_entities[s["admissionNo"]] = s["id"]

        parent_creds = []
        for i, (fn, ln) in enumerate(PARENT_DATA, 1):
            username = f"{fn.lower()}.{ln.lower()}{i}"
            password = f"{fn}@Parent{i}26!"
            email    = f"parent{i}@{tid}.edu"
            u = api("POST", "/users", token=adm_token, tenant=schema, json_body={
                "fullName": f"{fn} {ln}",
                "username": username,
                "email":    email,
                "password": password,
                "role":     "PARENT",
            })
            parent_user_id = None
            if ok(u):
                parent_user_id = u["data"]["id"]
                log("  ✅", f"{username}")
            elif "already" in str(u.get("message", "")).lower() or \
                 "duplicate" in str(u.get("message", "")).lower():
                log("  ℹ️", f"{username} (exists)")
            else:
                log("  ⚠️", f"{username}: {u.get('message')}")

            # Link parent to student i
            student_adm_no = f"{code}-S{i:03d}"
            student_id = stu_entities.get(student_adm_no)
            if parent_user_id and student_id:
                lk = api("POST", "/parents/links", token=adm_token, tenant=schema, json_body={
                    "parentUserId": parent_user_id,
                    "studentId":    student_id,
                })
                if ok(lk) or "already" in str(lk.get("message", "")).lower():
                    log("  ✅", f"Linked {username} → {student_adm_no}")
                else:
                    log("  ⚠️", f"Link failed ({username}): {lk.get('message')}")
            elif not student_id:
                log("  ⚠️", f"Student entity {student_adm_no} not found — link skipped")

            parent_creds.append({
                "name":      f"{fn} {ln}",
                "username":  username,
                "password":  password,
                "email":     email,
                "linkedTo":  student_adm_no,
                "role":      "PARENT",
            })

        all_credentials.append({
            "tenantId":    tid,
            "schoolName":  tenant["schoolName"],
            "schema":      schema,
            "admin":       adm,
            "teachers":    teacher_creds,
            "students":    student_creds[:5],
            "parents":     parent_creds,
        })

    # ── 4. Print credentials summary ──────────────────────────────
    print("\n" + "═"*60)
    print("  LOGIN CREDENTIALS SUMMARY")
    print("═"*60)
    print(f"\n{'SUPER ADMIN':}")
    print(f"  username : {SA_USER}")
    print(f"  password : {SA_PASS}")
    print(f"  login at : http://localhost:5173/super-admin/login")
    print(f"  NOTE     : Do NOT send X-Tenant-ID header")

    for t in all_credentials:
        tid    = t["tenantId"]
        schema = t["schema"]
        adm    = t["admin"]
        print(f"\n{'─'*60}")
        print(f"  {t['schoolName'].upper()}  (tenantId: {tid})")
        print(f"  X-Tenant-ID: {schema}")
        print(f"  login at  : http://localhost:5173/login")
        print(f"{'─'*60}")
        print(f"  [SCHOOL_ADMIN]")
        print(f"    username : {adm['username']}")
        print(f"    password : {adm['password']}")
        print()
        print(f"  [TEACHERS]")
        for tc in t["teachers"]:
            print(f"    {pad(tc['username'])} {pad(tc['password'])} ({tc['subject']})")
        print()
        print(f"  [STUDENTS — with login]")
        for sc in t["students"]:
            print(f"    {pad(sc['username'])} {pad(sc['password'])} adm:{sc['admNo']}")
        print()
        print(f"  [PARENTS — linked to students]")
        for pc in t.get("parents", []):
            print(f"    {pad(pc['username'])} {pad(pc['password'])} → {pc['linkedTo']}")

    # ── 5. Write JSON credentials file ────────────────────────────
    out_path = "/tmp/cloudcampus_credentials.json"
    with open(out_path, "w") as f:
        json.dump({
            "superAdmin": {"username": SA_USER, "password": SA_PASS},
            "tenants": all_credentials,
        }, f, indent=2)
    print(f"\n  📄 Full credentials saved to: {out_path}")
    print("\n" + "═"*60 + "\n")

if __name__ == "__main__":
    seed()

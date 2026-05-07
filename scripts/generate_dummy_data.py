#!/usr/bin/env python3
"""Generate CloudCampus dummy data Excel file."""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import os

TODAY = date.today()

# ── Colour palette ──────────────────────────────────────────────────────────
BLUE_DARK   = "1E3A5F"   # header bg
BLUE_MID    = "2563EB"   # accent
BLUE_LIGHT  = "DBEAFE"   # alt row
WHITE       = "FFFFFF"
GOLD        = "F59E0B"
GREEN       = "16A34A"
RED         = "DC2626"
GREY_LIGHT  = "F1F5F9"
GREY_TEXT   = "64748B"

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def alt_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_font(color=WHITE, bold=True, size=10):
    return Font(color=color, bold=bold, size=size, name="Calibri")

def cell_font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, name="Calibri", color=color)

def centered():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="CBD5E1")
thick = Side(style="medium", color="94A3B8")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_bottom():
    return Border(left=thin, right=thin, top=thin, bottom=thick)


def style_sheet_header(ws, title, subtitle=""):
    """Merge top rows for a sheet title banner."""
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    c = ws.cell(row=1, column=1, value=f"☁  CloudCampus — {title}")
    c.fill = hdr_fill(BLUE_DARK)
    c.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
    c.alignment = centered()
    if subtitle:
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.fill = hdr_fill(BLUE_MID)
        c2.font = Font(color=WHITE, size=9, name="Calibri", italic=True)
        c2.alignment = centered()


def write_table(ws, start_row, headers, rows,
                header_color=BLUE_MID, alt_color=BLUE_LIGHT):
    """Write a styled table starting at start_row."""
    # Header row
    ws.row_dimensions[start_row].height = 22
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=col_idx, value=h)
        c.fill = hdr_fill(header_color)
        c.font = hdr_font()
        c.alignment = centered()
        c.border = thick_bottom()

    # Data rows
    for r_idx, row in enumerate(rows, start=start_row + 1):
        ws.row_dimensions[r_idx].height = 18
        fill = alt_fill(alt_color) if r_idx % 2 == 0 else alt_fill(WHITE)
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.fill = fill
            c.font = cell_font()
            c.alignment = left_align()
            c.border = thin_border()

    return start_row + 1 + len(rows)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def merge_banner(ws, max_col, row, text, bg=BLUE_DARK):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=max_col)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = hdr_fill(bg)
    c.font = Font(color=WHITE, bold=True, size=10, name="Calibri")
    c.alignment = centered()
    ws.row_dimensions[row].height = 20


# ════════════════════════════════════════════════════════════════════════════
# Sheet builders
# ════════════════════════════════════════════════════════════════════════════

def build_overview(wb):
    ws = wb.create_sheet("📋 Overview", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    style_sheet_header(ws, "Demo Seed Data Overview")
    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1,
            value="School: Sunrise Academy  |  Tenant Slug: sunrise-academy  |  Default Password: Demo@2026!")
    ws.cell(row=2, column=1).fill = hdr_fill(BLUE_MID)
    ws.cell(row=2, column=1).font = Font(color=WHITE, size=10, name="Calibri")
    ws.cell(row=2, column=1).alignment = centered()

    headers = ["Sheet", "Description", "Records", "Username Format", "Password", "Tenant Slug"]
    rows = [
        ["👤 School Admin",    "School-level admin account",       "1",  "firstname.schooladmin", "Demo@2026!", "sunrise-academy"],
        ["👨‍🏫 Teachers",        "Teaching staff",                   "10", "firstname.teacher",     "Demo@2026!", "sunrise-academy"],
        ["🎓 Students",        "Enrolled students",                "15", "firstname.student",     "Demo@2026!", "sunrise-academy"],
        ["👨‍👩‍👦 Parents",          "Parent/guardian accounts",         "7",  "firstname.parent",      "Demo@2026!", "sunrise-academy"],
        ["🏫 Classes",         "Grade 1 – Grade 10",               "10", "—",                     "—",          "—"],
        ["📚 Sections",        "Sections per grade",               "18", "—",                     "—",          "—"],
        ["📖 Subjects",        "Curriculum subjects",              "12", "—",                     "—",          "—"],
        ["📝 Exams",           "Scheduled exams",                  "14", "—",                     "—",          "—"],
        ["📊 Exam Results",    "Published student results",        "29", "—",                     "—",          "—"],
        ["💰 Fees",            "Fee assignments + payments",       "24", "—",                     "—",          "—"],
        ["📓 Homework",        "Homework assignments",             "16", "—",                     "—",          "—"],
        ["🗓 Timetable",       "Weekly timetable slots",           "65+","—",                     "—",          "—"],
        ["✅ Attendance",      "30 days per student",              "450","—",                     "—",          "—"],
        ["🌐 Website CMS",     "Website config + sections",        "6",  "—",                     "—",          "—"],
        ["🖼 Gallery",         "Gallery images",                   "8",  "—",                     "—",          "—"],
        ["📬 Admission Leads", "Prospective student inquiries",    "12", "—",                     "—",          "—"],
    ]
    write_table(ws, 4, headers, rows, header_color=BLUE_DARK)
    set_col_widths(ws, [22, 34, 10, 28, 14, 18])

    # Login tip box
    ws.merge_cells("A23:F23")
    tip = ws.cell(row=23, column=1,
                  value='🔑  Login: POST /api/v1/auth/login  →  { "username": "...", "password": "Demo@2026!", "tenantSlug": "sunrise-academy" }')
    tip.fill = hdr_fill("FEF3C7")
    tip.font = Font(color="92400E", bold=True, size=10, name="Calibri")
    tip.alignment = centered()
    ws.row_dimensions[23].height = 22


def build_users(wb):
    ws = wb.create_sheet("👤 School Admin")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    style_sheet_header(ws, "School Admin",
                       "1 account — SCHOOL_ADMIN role — tenantSlug: sunrise-academy")

    headers = ["Username (role format)", "Internal DB Username", "Full Name",
               "Email", "Phone", "Role", "Password", "Tenant Slug"]
    rows = [
        ["priya.schooladmin", "priya.sharma", "Priya Sharma",
         "priya@sunrise.edu", "9000000001", "SCHOOL_ADMIN", "Demo@2026!", "sunrise-academy"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [24, 20, 18, 26, 14, 16, 14, 18])


def build_teachers(wb):
    ws = wb.create_sheet("👨‍🏫 Teachers")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    style_sheet_header(ws, "Teachers",
                       "10 accounts — TEACHER role — tenantSlug: sunrise-academy")

    headers = ["Username (role format)", "Internal DB Username", "Full Name",
               "Email", "Phone", "Employee No", "Hire Date", "Role", "Password", "Subjects Taught"]
    rows = [
        ["sunita.teacher",  "sunita.aggarwal",  "Sunita Aggarwal",  "sunita@sunrise.edu",  "9000000101", "SUN-T01", str(TODAY - timedelta(days=5*365)),  "TEACHER", "Demo@2026!", "Mathematics"],
        ["vikram.teacher",  "vikram.desai",     "Vikram Desai",     "vikram@sunrise.edu",  "9000000102", "SUN-T02", str(TODAY - timedelta(days=4*365)),  "TEACHER", "Demo@2026!", "Science, Biology"],
        ["lakshmi.teacher", "lakshmi.krishnan", "Lakshmi Krishnan", "lakshmi@sunrise.edu", "9000000103", "SUN-T03", str(TODAY - timedelta(days=6*365)),  "TEACHER", "Demo@2026!", "English"],
        ["rahul.teacher",   "rahul.mehta",      "Rahul Mehta",      "rahul@sunrise.edu",   "9000000104", "SUN-T04", str(TODAY - timedelta(days=3*365)),  "TEACHER", "Demo@2026!", "History, Chemistry"],
        ["asha.teacher",    "asha.nair",        "Asha Nair",        "asha@sunrise.edu",    "9000000105", "SUN-T05", str(TODAY - timedelta(days=7*365)),  "TEACHER", "Demo@2026!", "Hindi, Geography"],
        ["deepak.teacher",  "deepak.gupta",     "Deepak Gupta",     "deepak@sunrise.edu",  "9000000106", "SUN-T06", str(TODAY - timedelta(days=2*365)),  "TEACHER", "Demo@2026!", "Computer Science"],
        ["meena.teacher",   "meena.pillai",     "Meena Pillai",     "meena@sunrise.edu",   "9000000107", "SUN-T07", str(TODAY - timedelta(days=8*365)),  "TEACHER", "Demo@2026!", "Art & Craft"],
        ["arjun.teacher",   "arjun.verma",      "Arjun Verma",      "arjun@sunrise.edu",   "9000000108", "SUN-T08", str(TODAY - timedelta(days=1*365)),  "TEACHER", "Demo@2026!", "Physical Education"],
        ["pooja.teacher",   "pooja.iyer",       "Pooja Iyer",       "pooja@sunrise.edu",   "9000000109", "SUN-T09", str(TODAY - timedelta(days=4*365)),  "TEACHER", "Demo@2026!", "Hindi (Grade 5)"],
        ["suresh.teacher",  "suresh.rao",       "Suresh Rao",       "suresh@sunrise.edu",  "9000000110", "SUN-T10", str(TODAY - timedelta(days=10*365)), "TEACHER", "Demo@2026!", "Physics"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [22, 20, 18, 26, 14, 12, 13, 10, 12, 28])


def build_students(wb):
    ws = wb.create_sheet("🎓 Students")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    style_sheet_header(ws, "Students",
                       "15 accounts — STUDENT role — tenantSlug: sunrise-academy")

    headers = ["Username (role format)", "Internal DB Username", "Full Name",
               "Email", "Phone", "Admission No", "Date of Birth", "Gender", "Role", "Password"]
    rows = [
        ["aarav.student",     "aarav.sharma1",    "Aarav Sharma",    "aarav@sunrise.edu",           "9000000201", "SUN-S001", "2015-06-10", "MALE",   "STUDENT", "Demo@2026!"],
        ["diya.student",      "diya.patel2",      "Diya Patel",      "diya@sunrise.edu",            "9000000202", "SUN-S002", "2015-11-02", "FEMALE", "STUDENT", "Demo@2026!"],
        ["kabir.student",     "kabir.singh3",     "Kabir Singh",     "kabir@sunrise.edu",           "9000000203", "SUN-S003", "2014-03-22", "MALE",   "STUDENT", "Demo@2026!"],
        ["ananya.student",    "ananya.roy4",      "Ananya Roy",      "ananya@sunrise.edu",          "9000000204", "SUN-S004", "2014-07-15", "FEMALE", "STUDENT", "Demo@2026!"],
        ["rohan.student",     "rohan.kumar5",     "Rohan Kumar",     "rohan@sunrise.edu",           "9000000205", "SUN-S005", "2013-01-30", "MALE",   "STUDENT", "Demo@2026!"],
        ["priya.student",     "priya.joshi6",     "Priya Joshi",     "priyaj@sunrise.edu",          "9000000206", "SUN-S006", "2013-09-05", "FEMALE", "STUDENT", "Demo@2026!"],
        ["aryan.student",     "aryan.mehta7",     "Aryan Mehta",     "aryan@sunrise.edu",           "9000000207", "SUN-S007", "2012-04-18", "MALE",   "STUDENT", "Demo@2026!"],
        ["sneha.student",     "sneha.gupta8",     "Sneha Gupta",     "sneha@sunrise.edu",           "9000000208", "SUN-S008", "2012-08-27", "FEMALE", "STUDENT", "Demo@2026!"],
        ["vikrant.student",   "vikrant.nair9",    "Vikrant Nair",    "vikrant@sunrise.edu",         "9000000209", "SUN-S009", "2011-02-14", "MALE",   "STUDENT", "Demo@2026!"],
        ["riya.student",      "riya.iyer10",      "Riya Iyer",       "riya@sunrise.edu",            "9000000210", "SUN-S010", "2011-05-20", "FEMALE", "STUDENT", "Demo@2026!"],
        ["harsh.student",     "harsh.pandey11",   "Harsh Pandey",    "harsh@sunrise.edu",           "9000000211", "SUN-S011", "2010-06-03", "MALE",   "STUDENT", "Demo@2026!"],
        ["pooja.student",     "pooja.mishra12",   "Pooja Mishra",    "pooja.mishra@sunrise.edu",    "9000000212", "SUN-S012", "2010-10-11", "FEMALE", "STUDENT", "Demo@2026!"],
        ["aditya.student",    "aditya.das13",     "Aditya Das",      "aditya@sunrise.edu",          "9000000213", "SUN-S013", "2009-03-25", "MALE",   "STUDENT", "Demo@2026!"],
        ["kavya.student",     "kavya.menon14",    "Kavya Menon",     "kavya@sunrise.edu",           "9000000214", "SUN-S014", "2009-08-08", "FEMALE", "STUDENT", "Demo@2026!"],
        ["siddharth.student", "siddharth.rao15",  "Siddharth Rao",   "siddharth@sunrise.edu",       "9000000215", "SUN-S015", "2008-12-17", "MALE",   "STUDENT", "Demo@2026!"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [22, 20, 16, 28, 14, 12, 14, 8, 10, 12])


def build_parents(wb):
    ws = wb.create_sheet("👨‍👩‍👦 Parents")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    style_sheet_header(ws, "Parents",
                       "7 accounts — PARENT role — tenantSlug: sunrise-academy")

    headers = ["Username (role format)", "Internal DB Username", "Full Name",
               "Email", "Phone", "Role", "Password", "Linked Children"]
    rows = [
        ["ramesh.parent", "ramesh.sharma.p", "Ramesh Sharma", "ramesh.p@sunrise.edu", "9000000301", "PARENT", "Demo@2026!", "Aarav Sharma"],
        ["sujata.parent", "sujata.patel.p",  "Sujata Patel",  "sujata.p@sunrise.edu", "9000000302", "PARENT", "Demo@2026!", "Diya Patel"],
        ["mukesh.parent", "mukesh.singh.p",  "Mukesh Singh",  "mukesh.p@sunrise.edu", "9000000303", "PARENT", "Demo@2026!", "Kabir Singh, Ananya Roy"],
        ["geeta.parent",  "geeta.roy.p",     "Geeta Roy",     "geeta.p@sunrise.edu",  "9000000304", "PARENT", "Demo@2026!", "Rohan Kumar"],
        ["naveen.parent", "naveen.kumar.p",  "Naveen Kumar",  "naveen.p@sunrise.edu", "9000000305", "PARENT", "Demo@2026!", "Priya Joshi, Aryan Mehta"],
        ["anita.parent",  "anita.joshi.p",   "Anita Joshi",   "anita.p@sunrise.edu",  "9000000306", "PARENT", "Demo@2026!", "Sneha Gupta, Vikrant Nair"],
        ["rajesh.parent", "rajesh.mehta.p",  "Rajesh Mehta",  "rajesh.p@sunrise.edu", "9000000307", "PARENT", "Demo@2026!", "Riya Iyer"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [22, 20, 16, 26, 14, 10, 12, 32])


def build_classes(wb):
    ws = wb.create_sheet("🏫 Classes & Sections")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    style_sheet_header(ws, "Classes & Sections",
                       "10 grades · 18 sections (A/B/C)")

    headers = ["Class Name", "Code", "Sections", "Total Sections", "Active"]
    rows = [
        ["Grade 1",  "G01", "A, B, C", "3", "Yes"],
        ["Grade 2",  "G02", "A, B",    "2", "Yes"],
        ["Grade 3",  "G03", "A, B, C", "3", "Yes"],
        ["Grade 4",  "G04", "A, B",    "2", "Yes"],
        ["Grade 5",  "G05", "A, B",    "2", "Yes"],
        ["Grade 6",  "G06", "A, B",    "2", "Yes"],
        ["Grade 7",  "G07", "A",       "1", "Yes"],
        ["Grade 8",  "G08", "A, B",    "2", "Yes"],
        ["Grade 9",  "G09", "A",       "1", "Yes"],
        ["Grade 10", "G10", "A",       "1", "Yes"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [14, 10, 18, 16, 8])


def build_subjects(wb):
    ws = wb.create_sheet("📖 Subjects")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    ws.merge_cells("A2:C2")
    style_sheet_header(ws, "Subjects", "12 curriculum subjects")

    headers = ["Subject Name", "Code", "Active"]
    rows = [
        ["Mathematics",        "MATH", "Yes"],
        ["Science",            "SCI",  "Yes"],
        ["English",            "ENG",  "Yes"],
        ["Hindi",              "HIN",  "Yes"],
        ["History",            "HIST", "Yes"],
        ["Geography",          "GEO",  "Yes"],
        ["Physics",            "PHY",  "Yes"],
        ["Chemistry",          "CHEM", "Yes"],
        ["Biology",            "BIO",  "Yes"],
        ["Computer Science",   "CS",   "Yes"],
        ["Art & Craft",        "ART",  "Yes"],
        ["Physical Education", "PE",   "Yes"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [24, 10, 8])


def build_exams(wb):
    ws = wb.create_sheet("📝 Exams")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    style_sheet_header(ws, "Exams", "14 scheduled exams across Grades 1, 3, 5 & 10")

    headers = ["Exam Title", "Class", "Section", "Subject",
               "Teacher (role format)", "Max Marks", "Exam Date", "Active"]
    rows = [
        ["Unit Test 1 - Mathematics",  "Grade 1",  "A", "Mathematics",      "sunita.teacher",  "100", str(TODAY - timedelta(days=60)), "Yes"],
        ["Unit Test 1 - English",      "Grade 1",  "A", "English",          "lakshmi.teacher", "100", str(TODAY - timedelta(days=58)), "Yes"],
        ["Mid-Term - Mathematics",     "Grade 3",  "A", "Mathematics",      "sunita.teacher",  "100", str(TODAY - timedelta(days=45)), "Yes"],
        ["Mid-Term - Science",         "Grade 3",  "A", "Science",          "vikram.teacher",  "100", str(TODAY - timedelta(days=44)), "Yes"],
        ["Mid-Term - English",         "Grade 3",  "A", "English",          "lakshmi.teacher", "100", str(TODAY - timedelta(days=43)), "Yes"],
        ["Final Exam - Mathematics",   "Grade 5",  "A", "Mathematics",      "sunita.teacher",  "100", str(TODAY - timedelta(days=15)), "Yes"],
        ["Final Exam - Science",       "Grade 5",  "A", "Science",          "vikram.teacher",  "100", str(TODAY - timedelta(days=14)), "Yes"],
        ["Final Exam - History",       "Grade 5",  "A", "History",          "rahul.teacher",   "100", str(TODAY - timedelta(days=13)), "Yes"],
        ["Board Mock - Physics",       "Grade 10", "A", "Physics",          "suresh.teacher",  "100", str(TODAY - timedelta(days=30)), "Yes"],
        ["Board Mock - Chemistry",     "Grade 10", "A", "Chemistry",        "rahul.teacher",   "100", str(TODAY - timedelta(days=29)), "Yes"],
        ["Board Mock - Mathematics",   "Grade 10", "A", "Mathematics",      "sunita.teacher",  "100", str(TODAY - timedelta(days=28)), "Yes"],
        ["Board Mock - Biology",       "Grade 10", "A", "Biology",          "vikram.teacher",  "100", str(TODAY - timedelta(days=27)), "Yes"],
        ["Unit Test 2 - Mathematics",  "Grade 1",  "B", "Mathematics",      "sunita.teacher",  "50",  str(TODAY - timedelta(days=20)), "Yes"],
        ["Unit Test 2 - Hindi",        "Grade 3",  "B", "Hindi",            "asha.teacher",    "50",  str(TODAY - timedelta(days=19)), "Yes"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [30, 10, 9, 14, 20, 11, 13, 8])


def build_exam_results(wb):
    ws = wb.create_sheet("📊 Exam Results")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")
    style_sheet_header(ws, "Exam Results", "29 published results")

    headers = ["Exam Title", "Student (role format)", "Student Full Name",
               "Marks Obtained", "Max Marks", "Grade", "Remarks"]
    rows = [
        ["Unit Test 1 - Mathematics", "aarav.student",     "Aarav Sharma",   "85", "100", "B+", "Good effort"],
        ["Unit Test 1 - Mathematics", "diya.student",      "Diya Patel",     "92", "100", "A",  "Excellent"],
        ["Unit Test 1 - English",     "aarav.student",     "Aarav Sharma",   "78", "100", "B",  "Needs improvement in grammar"],
        ["Unit Test 1 - English",     "diya.student",      "Diya Patel",     "88", "100", "A-", "Very good"],
        ["Mid-Term - Mathematics",    "kabir.student",     "Kabir Singh",    "74", "100", "B-", "Satisfactory"],
        ["Mid-Term - Mathematics",    "ananya.student",    "Ananya Roy",     "90", "100", "A",  "Outstanding"],
        ["Mid-Term - Science",        "kabir.student",     "Kabir Singh",    "68", "100", "C+", "Average"],
        ["Mid-Term - Science",        "ananya.student",    "Ananya Roy",     "95", "100", "A+", "Brilliant"],
        ["Mid-Term - English",        "kabir.student",     "Kabir Singh",    "80", "100", "B",  "Good"],
        ["Mid-Term - English",        "ananya.student",    "Ananya Roy",     "72", "100", "B-", "Keep trying"],
        ["Final Exam - Mathematics",  "rohan.student",     "Rohan Kumar",    "88", "100", "A-", "Well done"],
        ["Final Exam - Mathematics",  "priya.student",     "Priya Joshi",    "65", "100", "C",  "Need more practice"],
        ["Final Exam - Science",      "rohan.student",     "Rohan Kumar",    "91", "100", "A",  "Excellent"],
        ["Final Exam - Science",      "priya.student",     "Priya Joshi",    "70", "100", "B-", "Good attempt"],
        ["Final Exam - History",      "rohan.student",     "Rohan Kumar",    "76", "100", "B",  "Good"],
        ["Final Exam - History",      "priya.student",     "Priya Joshi",    "82", "100", "B+", "Very good"],
        ["Board Mock - Physics",      "aditya.student",    "Aditya Das",     "78", "100", "B",  "Good"],
        ["Board Mock - Physics",      "kavya.student",     "Kavya Menon",    "84", "100", "B+", "Very good"],
        ["Board Mock - Physics",      "siddharth.student", "Siddharth Rao",  "91", "100", "A",  "Excellent"],
        ["Board Mock - Chemistry",    "aditya.student",    "Aditya Das",     "72", "100", "B-", "Satisfactory"],
        ["Board Mock - Chemistry",    "kavya.student",     "Kavya Menon",    "88", "100", "A-", "Outstanding"],
        ["Board Mock - Chemistry",    "siddharth.student", "Siddharth Rao",  "95", "100", "A+", "Perfect"],
        ["Board Mock - Mathematics",  "aditya.student",    "Aditya Das",     "80", "100", "B",  "Good"],
        ["Board Mock - Mathematics",  "kavya.student",     "Kavya Menon",    "76", "100", "B",  "Keep it up"],
        ["Board Mock - Mathematics",  "siddharth.student", "Siddharth Rao",  "98", "100", "A+", "Brilliant"],
        ["Board Mock - Biology",      "aditya.student",    "Aditya Das",     "69", "100", "C+", "Average"],
        ["Board Mock - Biology",      "kavya.student",     "Kavya Menon",    "92", "100", "A",  "Excellent"],
        ["Board Mock - Biology",      "siddharth.student", "Siddharth Rao",  "87", "100", "A-", "Very good"],
        ["Unit Test 2 - Mathematics", "aarav.student",     "Aarav Sharma",   "44", "50",  "A",  "Well done"],
        ["Unit Test 2 - Hindi",       "kabir.student",     "Kabir Singh",    "38", "50",  "B+", "Good"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [30, 22, 18, 15, 11, 8, 30])


def build_fees(wb):
    ws = wb.create_sheet("💰 Fees")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:I1")
    ws.merge_cells("A2:I2")
    style_sheet_header(ws, "Fee Assignments & Payments",
                       "24 fee assignments — various statuses and payment methods")

    headers = ["Student (role format)", "Student Full Name", "Fee Title",
               "Amount (₹)", "Due Date", "Status", "Paid Amount (₹)", "Payment Method", "Reference No"]

    def due(days): return str(TODAY - timedelta(days=days)) if days > 0 else str(TODAY + timedelta(days=abs(days)))

    rows = [
        ["aarav.student",     "Aarav Sharma",   "Term 1 Tuition Fee",  "15000", due(90),  "PAID",           "15000", "BANK_TRANSFER", "RCP-001"],
        ["aarav.student",     "Aarav Sharma",   "Activity Fee",         "2500", due(60),  "PAID",            "2500", "CASH",          "—"],
        ["aarav.student",     "Aarav Sharma",   "Term 2 Tuition Fee",  "15000", due(-30), "PENDING",            "0", "—",             "—"],
        ["diya.student",      "Diya Patel",     "Term 1 Tuition Fee",  "15000", due(90),  "PAID",           "15000", "CASH",          "—"],
        ["diya.student",      "Diya Patel",     "Exam Fee",             "1500", due(30),  "PAID",            "1500", "BANK_TRANSFER", "RCP-002"],
        ["diya.student",      "Diya Patel",     "Term 2 Tuition Fee",  "15000", due(-30), "PENDING",            "0", "—",             "—"],
        ["kabir.student",     "Kabir Singh",    "Term 1 Tuition Fee",  "18000", due(90),  "PARTIALLY_PAID",  "9000", "CASH",          "—"],
        ["kabir.student",     "Kabir Singh",    "Library Fee",           "500", due(60),  "PAID",             "500", "CASH",          "—"],
        ["ananya.student",    "Ananya Roy",     "Term 1 Tuition Fee",  "18000", due(90),  "PAID",           "18000", "CHEQUE",        "CHQ-301"],
        ["ananya.student",    "Ananya Roy",     "Sports Fee",           "1000", due(45),  "PAID",            "1000", "CASH",          "—"],
        ["rohan.student",     "Rohan Kumar",    "Term 1 Tuition Fee",  "20000", due(90),  "PAID",           "20000", "BANK_TRANSFER", "RCP-003"],
        ["rohan.student",     "Rohan Kumar",    "Term 2 Tuition Fee",  "20000", due(-20), "PENDING",            "0", "—",             "—"],
        ["rohan.student",     "Rohan Kumar",    "Exam Fee",             "2000", due(30),  "PAID",            "2000", "CASH",          "—"],
        ["priya.student",     "Priya Joshi",    "Term 1 Tuition Fee",  "20000", due(90),  "PARTIALLY_PAID", "10000", "CASH",          "—"],
        ["aryan.student",     "Aryan Mehta",    "Term 1 Tuition Fee",  "22000", due(90),  "PAID",           "22000", "BANK_TRANSFER", "RCP-004"],
        ["aryan.student",     "Aryan Mehta",    "Lab Fee",              "1500", due(30),  "PAID",            "1500", "CASH",          "—"],
        ["sneha.student",     "Sneha Gupta",    "Term 1 Tuition Fee",  "22000", due(90),  "PAID",           "22000", "CHEQUE",        "CHQ-302"],
        ["vikrant.student",   "Vikrant Nair",   "Term 1 Tuition Fee",  "25000", due(90),  "PENDING",            "0", "—",             "—"],
        ["riya.student",      "Riya Iyer",      "Term 1 Tuition Fee",  "25000", due(90),  "PAID",           "25000", "BANK_TRANSFER", "RCP-005"],
        ["harsh.student",     "Harsh Pandey",   "Term 1 Tuition Fee",  "28000", due(90),  "PAID",           "28000", "BANK_TRANSFER", "RCP-006"],
        ["pooja.student",     "Pooja Mishra",   "Term 1 Tuition Fee",  "28000", due(90),  "PARTIALLY_PAID", "14000", "CASH",          "—"],
        ["aditya.student",    "Aditya Das",     "Term 1 Tuition Fee",  "28000", due(90),  "PAID",           "28000", "CHEQUE",        "CHQ-303"],
        ["kavya.student",     "Kavya Menon",    "Term 1 Tuition Fee",  "28000", due(90),  "PAID",           "28000", "BANK_TRANSFER", "RCP-007"],
        ["siddharth.student", "Siddharth Rao",  "Term 1 Tuition Fee",  "28000", due(90),  "PAID",           "28000", "BANK_TRANSFER", "RCP-008"],
        ["siddharth.student", "Siddharth Rao",  "Board Exam Fee",       "3000", due(20),  "PAID",            "3000", "BANK_TRANSFER", "RCP-009"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [22, 16, 22, 13, 14, 18, 16, 16, 12])

    # Colour-code status column (col 6 = F)
    for row_idx in range(5, 5 + len(rows)):
        cell = ws.cell(row=row_idx, column=6)
        if cell.value == "PAID":
            cell.font = Font(color=GREEN, bold=True, size=10, name="Calibri")
        elif cell.value == "PENDING":
            cell.font = Font(color=RED, bold=True, size=10, name="Calibri")
        elif cell.value == "PARTIALLY_PAID":
            cell.font = Font(color=GOLD, bold=True, size=10, name="Calibri")


def build_homework(wb):
    ws = wb.create_sheet("📓 Homework")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")
    style_sheet_header(ws, "Homework Assignments", "16 assignments across Grades 1, 3, 5 & 10")

    headers = ["Title", "Instructions (summary)", "Class", "Section",
               "Teacher (role format)", "Due Date", "Active"]

    def fwd(days): return str(TODAY + timedelta(days=days))

    rows = [
        ["Chapter 2 – Place Value Practice",    "Complete exercises 2.1–2.5 from textbook",        "Grade 1",  "A", "sunita.teacher",  fwd(3), "Yes"],
        ["Grammar Exercise – Tenses",           "Fill blanks with correct verb forms",              "Grade 1",  "A", "lakshmi.teacher", fwd(4), "Yes"],
        ["Draw Animals Using Shapes",           "Draw 5 animals using basic geometric shapes",      "Grade 1",  "B", "meena.teacher",   fwd(2), "Yes"],
        ["Fractions Worksheet",                 "Solve 20 fraction problems from worksheet",        "Grade 3",  "A", "sunita.teacher",  fwd(3), "Yes"],
        ["Plant Cell Diagram",                  "Draw and label a plant cell with 10 parts",        "Grade 3",  "A", "vikram.teacher",  fwd(5), "Yes"],
        ["Essay – My Favourite Season",         "Write a 200-word essay on your favourite season", "Grade 3",  "A", "lakshmi.teacher", fwd(4), "Yes"],
        ["Hindi Paragraph Writing",             "Write 5 sentences about your family in Hindi",    "Grade 3",  "B", "asha.teacher",    fwd(2), "Yes"],
        ["Algebraic Expressions Practice",      "Simplify expressions – problems 1–30",             "Grade 5",  "A", "sunita.teacher",  fwd(3), "Yes"],
        ["Water Cycle Project",                 "Make a poster showing water cycle steps",          "Grade 5",  "A", "vikram.teacher",  fwd(7), "Yes"],
        ["Medieval India – Summary",            "Write a 300-word summary of Chapter 4",           "Grade 5",  "A", "rahul.teacher",   fwd(4), "Yes"],
        ["Python Turtle Program",               "Write program to draw a star using turtle",        "Grade 5",  "B", "deepak.teacher",  fwd(5), "Yes"],
        ["Optics Problems Set",                 "Solve numerical problems 1–15 on reflection",     "Grade 10", "A", "suresh.teacher",  fwd(3), "Yes"],
        ["Organic Chemistry Reactions",         "Balance 10 organic reaction equations",           "Grade 10", "A", "rahul.teacher",   fwd(4), "Yes"],
        ["Trigonometry – Prove 15 Identities",  "Prove identities from exercise 8.3",              "Grade 10", "A", "sunita.teacher",  fwd(5), "Yes"],
        ["Database Design Assignment",          "Design an ER diagram for a library system",       "Grade 10", "A", "deepak.teacher",  fwd(6), "Yes"],
        ["Photorespiration Essay",              "Write 400 words on photorespiration",             "Grade 10", "A", "vikram.teacher",  fwd(3), "Yes"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [36, 42, 10, 9, 20, 13, 8])


def build_timetable(wb):
    ws = wb.create_sheet("🗓 Timetable")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    style_sheet_header(ws, "Timetable Slots",
                       "65+ slots — Grade 1-A (Mon–Fri), Grade 5-A (Mon–Fri), Grade 10-A (Mon–Fri)")

    headers = ["Class", "Section", "Day", "Day No",
               "Start Time", "End Time", "Subject", "Teacher (role format)"]

    DAY = {1: "Monday", 2: "Tuesday", 3: "Wednesday", 4: "Thursday", 5: "Friday"}

    rows = []
    # Grade 1-A
    slots_g1a = [
        (1,"07:30","08:30","Mathematics","sunita.teacher"),
        (1,"08:30","09:30","English","lakshmi.teacher"),
        (1,"09:45","10:45","Hindi","asha.teacher"),
        (1,"10:45","11:45","Science","vikram.teacher"),
        (1,"12:30","13:30","Art & Craft","meena.teacher"),
        (1,"13:30","14:30","Physical Education","arjun.teacher"),
        (2,"07:30","08:30","English","lakshmi.teacher"),
        (2,"08:30","09:30","Mathematics","sunita.teacher"),
        (2,"09:45","10:45","Science","vikram.teacher"),
        (2,"10:45","11:45","Hindi","asha.teacher"),
        (2,"12:30","13:30","Art & Craft","meena.teacher"),
        (3,"07:30","08:30","Mathematics","sunita.teacher"),
        (3,"08:30","09:30","Science","vikram.teacher"),
        (3,"09:45","10:45","English","lakshmi.teacher"),
        (4,"07:30","08:30","Mathematics","sunita.teacher"),
        (4,"08:30","09:30","Hindi","asha.teacher"),
        (4,"09:45","10:45","Science","vikram.teacher"),
        (5,"07:30","08:30","English","lakshmi.teacher"),
        (5,"08:30","09:30","Mathematics","sunita.teacher"),
        (5,"09:45","10:45","Physical Education","arjun.teacher"),
    ]
    for d,s,e,sub,t in slots_g1a:
        rows.append(["Grade 1", "A", DAY[d], str(d), s, e, sub, t])

    # Grade 5-A
    slots_g5a = [
        (1,"07:30","08:30","Mathematics","sunita.teacher"),
        (1,"08:30","09:30","Science","vikram.teacher"),
        (1,"09:45","10:45","English","lakshmi.teacher"),
        (1,"10:45","11:45","History","rahul.teacher"),
        (2,"07:30","08:30","Geography","asha.teacher"),
        (2,"08:30","09:30","Mathematics","sunita.teacher"),
        (2,"09:45","10:45","Hindi","pooja.teacher"),
        (3,"07:30","08:30","Computer Science","deepak.teacher"),
        (3,"08:30","09:30","Mathematics","sunita.teacher"),
        (3,"09:45","10:45","Science","vikram.teacher"),
        (4,"07:30","08:30","English","lakshmi.teacher"),
        (4,"08:30","09:30","History","rahul.teacher"),
        (4,"09:45","10:45","Physical Education","arjun.teacher"),
        (5,"07:30","08:30","Mathematics","sunita.teacher"),
        (5,"08:30","09:30","Geography","asha.teacher"),
        (5,"09:45","10:45","Computer Science","deepak.teacher"),
    ]
    for d,s,e,sub,t in slots_g5a:
        rows.append(["Grade 5", "A", DAY[d], str(d), s, e, sub, t])

    # Grade 10-A
    slots_g10a = [
        (1,"07:30","08:30","Mathematics","sunita.teacher"),
        (1,"08:30","09:30","Physics","suresh.teacher"),
        (1,"09:45","10:45","Chemistry","rahul.teacher"),
        (1,"10:45","11:45","Biology","vikram.teacher"),
        (2,"07:30","08:30","English","lakshmi.teacher"),
        (2,"08:30","09:30","Mathematics","sunita.teacher"),
        (2,"09:45","10:45","Computer Science","deepak.teacher"),
        (3,"07:30","08:30","Physics","suresh.teacher"),
        (3,"08:30","09:30","Chemistry","rahul.teacher"),
        (3,"09:45","10:45","Mathematics","sunita.teacher"),
        (4,"07:30","08:30","Biology","vikram.teacher"),
        (4,"08:30","09:30","English","lakshmi.teacher"),
        (5,"07:30","08:30","Computer Science","deepak.teacher"),
        (5,"08:30","09:30","Physics","suresh.teacher"),
        (5,"09:45","10:45","Physical Education","arjun.teacher"),
    ]
    for d,s,e,sub,t in slots_g10a:
        rows.append(["Grade 10", "A", DAY[d], str(d), s, e, sub, t])

    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [10, 9, 12, 8, 12, 12, 22, 22])


def build_attendance(wb):
    ws = wb.create_sheet("✅ Attendance")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    style_sheet_header(ws, "Attendance",
                       "30 school days × 15 students — weekdays only, going back from today")

    cycle = ["PRESENT","PRESENT","PRESENT","PRESENT","LATE",
             "PRESENT","PRESENT","ABSENT","PRESENT","PRESENT"]
    students = [
        ("aarav.student",     "Aarav Sharma",    "Grade 1",  "A"),
        ("diya.student",      "Diya Patel",      "Grade 1",  "A"),
        ("kabir.student",     "Kabir Singh",     "Grade 3",  "A"),
        ("ananya.student",    "Ananya Roy",      "Grade 3",  "A"),
        ("rohan.student",     "Rohan Kumar",     "Grade 5",  "A"),
        ("priya.student",     "Priya Joshi",     "Grade 5",  "A"),
        ("aryan.student",     "Aryan Mehta",     "Grade 7",  "A"),
        ("sneha.student",     "Sneha Gupta",     "Grade 7",  "A"),
        ("vikrant.student",   "Vikrant Nair",    "Grade 9",  "A"),
        ("riya.student",      "Riya Iyer",       "Grade 9",  "A"),
        ("harsh.student",     "Harsh Pandey",    "Grade 10", "A"),
        ("pooja.student",     "Pooja Mishra",    "Grade 10", "A"),
        ("aditya.student",    "Aditya Das",      "Grade 10", "A"),
        ("kavya.student",     "Kavya Menon",     "Grade 10", "A"),
        ("siddharth.student", "Siddharth Rao",   "Grade 10", "A"),
    ]

    headers = ["Student (role format)", "Student Name", "Class", "Section",
               "Date", "Status"]
    rows = []
    for uname, fullname, cls, sec in students:
        day_offset = 0
        recorded = 0
        while recorded < 30:
            d = TODAY - timedelta(days=day_offset + 1)
            day_offset += 1
            if d.weekday() >= 5:
                continue
            rows.append([uname, fullname, cls, sec,
                         str(d), cycle[recorded % len(cycle)]])
            recorded += 1

    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [22, 18, 10, 9, 14, 12])

    # Colour-code status
    for row_idx in range(5, 5 + len(rows)):
        cell = ws.cell(row=row_idx, column=6)
        if cell.value == "PRESENT":
            cell.font = Font(color=GREEN, bold=False, size=10, name="Calibri")
        elif cell.value == "ABSENT":
            cell.font = Font(color=RED, bold=True, size=10, name="Calibri")
        elif cell.value == "LATE":
            cell.font = Font(color=GOLD, bold=True, size=10, name="Calibri")


def build_website(wb):
    ws = wb.create_sheet("🌐 Website CMS")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    ws.merge_cells("A2:C2")
    style_sheet_header(ws, "Website CMS — Config, Sections & Gallery",
                       "Public schema — tenantId: sunrise-academy")

    # Config
    merge_banner(ws, 3, 4, "WEBSITE CONFIG", BLUE_DARK)
    config_headers = ["Property", "Value", "Notes"]
    config_rows = [
        ["tenantId",        "sunrise-academy",                     "Matches tenant slug"],
        ["Tagline",         "Enlightening Minds, Building Futures", ""],
        ["Email",           "info@sunriseacademy.edu.in",           ""],
        ["Phone",           "+91 98765 43210",                      ""],
        ["Address",         "12, Knowledge Park, Sector 18",        ""],
        ["City",            "Noida",                                ""],
        ["State",           "Uttar Pradesh",                        ""],
        ["Pincode",         "201301",                               ""],
        ["Theme Color",     "#2563EB",                              "Blue"],
        ["Admissions Open", "Yes",                                  ""],
        ["Facebook",        "https://facebook.com/sunriseacademy",  ""],
        ["Twitter",         "https://twitter.com/sunriseacademy",   ""],
        ["Instagram",       "https://instagram.com/sunriseacademy", ""],
        ["YouTube",         "https://youtube.com/sunriseacademy",   ""],
    ]
    next_row = write_table(ws, 5, config_headers, config_rows) + 1

    # Sections
    merge_banner(ws, 3, next_row, "WEBSITE SECTIONS (6)", BLUE_MID)
    sec_headers = ["Section Key", "Title", "Subtitle / Notes"]
    sec_rows = [
        ["hero",         "Welcome to Sunrise Academy",   "CTA: Apply Now"],
        ["about",        "About Us",                     "Founded 1995 · CBSE · 2000+ students · 120+ staff"],
        ["features",     "Why Choose Sunrise?",          "6 feature cards: Classrooms, Sports, Labs, Computers, Library, Arts"],
        ["faculty",      "Our Faculty",                  "Avg 12 yrs exp · 8 PhDs · 3 national award winners"],
        ["achievements", "Our Achievements",             "5 state toppers · 42 medals · 18 olympiad qualifiers"],
        ["contact",      "Get in Touch",                 "Office hours: Mon–Sat 8AM–4PM"],
    ]
    next_row = write_table(ws, next_row + 1, sec_headers, sec_rows) + 1

    # Gallery
    merge_banner(ws, 3, next_row, "GALLERY ITEMS (8)", BLUE_MID)
    gal_headers = ["Display Order", "Caption", "Image URL"]
    gal_rows = [
        ["0", "Annual Science Exhibition 2025",       "https://images.unsplash.com/photo-1580582932707"],
        ["1", "Inter-School Cricket Championship",    "https://images.unsplash.com/photo-1546410531-bb4"],
        ["2", "Grade 10 Board Felicitation Ceremony", "https://images.unsplash.com/photo-1509062522246"],
        ["3", "Annual Day Dance Performance",         "https://images.unsplash.com/photo-1604881991720"],
        ["4", "Robotics Club Workshop",               "https://images.unsplash.com/photo-1571260899304"],
        ["5", "Library Reading Programme",            "https://images.unsplash.com/photo-1503676260728"],
        ["6", "Art Exhibition – Student Masterpieces","https://images.unsplash.com/photo-1544717305-27"],
        ["7", "Smart Classroom Inauguration",         "https://images.unsplash.com/photo-1516979187457"],
    ]
    write_table(ws, next_row + 1, gal_headers, gal_rows)
    set_col_widths(ws, [28, 40, 50])


def build_admission_leads(wb):
    ws = wb.create_sheet("📬 Admission Leads")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")
    style_sheet_header(ws, "Admission Leads", "12 enquiries — public schema")

    headers = ["Parent Name", "Parent Email", "Parent Phone",
               "Student Name", "Applying Class", "Message", "Status"]
    rows = [
        ["Anil Kapoor",   "anil.kapoor@gmail.com",   "9811001001", "Ravi Kapoor",    "Grade 3",  "Interested. Please share prospectus.",                  "NEW"],
        ["Sunita Verma",  "sunita.v@yahoo.com",      "9811001002", "Priti Verma",    "Grade 1",  "Looking for CBSE school. Is transport available?",      "CONTACTED"],
        ["Mohan Das",     "mohan.das@outlook.com",   "9811001003", "Arjun Das",      "Grade 6",  "My son is a state-level chess player.",                 "NEW"],
        ["Lakshmi Nair",  "lakshmi.nair@gmail.com",  "9811001004", "Meera Nair",     "Grade 2",  "Heard great reviews. Would like to visit campus.",      "VISITED"],
        ["Prasad Reddy",  "prasad.r@gmail.com",      "9811001005", "Vignesh Reddy",  "Grade 9",  "Transferring from Hyderabad. Need mid-year admission.", "NEW"],
        ["Kiran Sharma",  "kiran.s@hotmail.com",     "9811001006", "Nisha Sharma",   "Grade 4",  "Looking for strong science programme.",                 "CONTACTED"],
        ["Deepa Iyer",    "deepa.iyer@gmail.com",    "9811001007", "Gautam Iyer",    "Grade 7",  "Interested. Please call to schedule a visit.",          "VISITED"],
        ["Rajiv Khanna",  "rajiv.khanna@gmail.com",  "9811001008", "Aakash Khanna",  "Grade 1",  "Just moved to Noida. Need admission urgently.",         "ADMITTED"],
        ["Preethi Menon", "preethi.m@gmail.com",     "9811001009", "Divya Menon",    "Grade 5",  "My daughter loves art and music.",                      "NEW"],
        ["Sachin Gupta",  "sachin.g@gmail.com",      "9811001010", "Ronak Gupta",    "Grade 8",  "Is there a scholarship for merit students?",            "CONTACTED"],
        ["Ananya Pillai", "ananya.p@gmail.com",      "9811001011", "Tara Pillai",    "Grade 3",  "Very interested. Visited campus last week.",            "VISITED"],
        ["Vijay Mishra",  "vijay.m@gmail.com",       "9811001012", "Shrey Mishra",   "Grade 10", "Need TC from current school. Can admission happen?",    "NEW"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [16, 26, 14, 16, 14, 44, 12])

    # Status colour
    status_colors = {
        "NEW": BLUE_MID, "CONTACTED": GOLD, "VISITED": GREEN, "ADMITTED": "16A34A"
    }
    for row_idx in range(5, 5 + len(rows)):
        cell = ws.cell(row=row_idx, column=7)
        color = status_colors.get(cell.value, "000000")
        cell.font = Font(color=color, bold=True, size=10, name="Calibri")


def build_quick_login(wb):
    ws = wb.create_sheet("🔑 Quick Login")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    style_sheet_header(ws, "Quick Login Reference",
                       "Copy-paste ready credentials for every role")

    headers = ["Role", "Username (role format)", "Password", "Tenant Slug", "Full Name", "Notes"]
    rows = [
        ["SUPER_ADMIN",  "<BOOTSTRAP_ADMIN_USERNAME>", "<BOOTSTRAP_ADMIN_PASSWORD>", "— (not needed)", "Platform Admin",  "Set via env vars"],
        ["SCHOOL_ADMIN", "priya.schooladmin",           "Demo@2026!",                "sunrise-academy", "Priya Sharma",    "Full school access"],
        ["TEACHER",      "sunita.teacher",              "Demo@2026!",                "sunrise-academy", "Sunita Aggarwal", "Mathematics"],
        ["TEACHER",      "vikram.teacher",              "Demo@2026!",                "sunrise-academy", "Vikram Desai",    "Science & Biology"],
        ["TEACHER",      "lakshmi.teacher",             "Demo@2026!",                "sunrise-academy", "Lakshmi Krishnan","English"],
        ["TEACHER",      "rahul.teacher",               "Demo@2026!",                "sunrise-academy", "Rahul Mehta",     "History & Chemistry"],
        ["TEACHER",      "asha.teacher",                "Demo@2026!",                "sunrise-academy", "Asha Nair",       "Hindi & Geography"],
        ["TEACHER",      "deepak.teacher",              "Demo@2026!",                "sunrise-academy", "Deepak Gupta",    "Computer Science"],
        ["TEACHER",      "meena.teacher",               "Demo@2026!",                "sunrise-academy", "Meena Pillai",    "Art & Craft"],
        ["TEACHER",      "arjun.teacher",               "Demo@2026!",                "sunrise-academy", "Arjun Verma",     "Physical Education"],
        ["TEACHER",      "pooja.teacher",               "Demo@2026!",                "sunrise-academy", "Pooja Iyer",      "Hindi (Grade 5)"],
        ["TEACHER",      "suresh.teacher",              "Demo@2026!",                "sunrise-academy", "Suresh Rao",      "Physics"],
        ["STUDENT",      "aarav.student",               "Demo@2026!",                "sunrise-academy", "Aarav Sharma",    "Grade 1-A · SUN-S001"],
        ["STUDENT",      "diya.student",                "Demo@2026!",                "sunrise-academy", "Diya Patel",      "Grade 1-A · SUN-S002"],
        ["STUDENT",      "kabir.student",               "Demo@2026!",                "sunrise-academy", "Kabir Singh",     "Grade 3-A · SUN-S003"],
        ["STUDENT",      "ananya.student",              "Demo@2026!",                "sunrise-academy", "Ananya Roy",      "Grade 3-A · SUN-S004"],
        ["STUDENT",      "rohan.student",               "Demo@2026!",                "sunrise-academy", "Rohan Kumar",     "Grade 5-A · SUN-S005"],
        ["STUDENT",      "priya.student",               "Demo@2026!",                "sunrise-academy", "Priya Joshi",     "Grade 5-A · SUN-S006"],
        ["STUDENT",      "aryan.student",               "Demo@2026!",                "sunrise-academy", "Aryan Mehta",     "Grade 7-A · SUN-S007"],
        ["STUDENT",      "sneha.student",               "Demo@2026!",                "sunrise-academy", "Sneha Gupta",     "Grade 7-A · SUN-S008"],
        ["STUDENT",      "vikrant.student",             "Demo@2026!",                "sunrise-academy", "Vikrant Nair",    "Grade 9-A · SUN-S009"],
        ["STUDENT",      "riya.student",                "Demo@2026!",                "sunrise-academy", "Riya Iyer",       "Grade 9-A · SUN-S010"],
        ["STUDENT",      "harsh.student",               "Demo@2026!",                "sunrise-academy", "Harsh Pandey",    "Grade 10-A · SUN-S011"],
        ["STUDENT",      "pooja.student",               "Demo@2026!",                "sunrise-academy", "Pooja Mishra",    "Grade 10-A · SUN-S012"],
        ["STUDENT",      "aditya.student",              "Demo@2026!",                "sunrise-academy", "Aditya Das",      "Grade 10-A · SUN-S013"],
        ["STUDENT",      "kavya.student",               "Demo@2026!",                "sunrise-academy", "Kavya Menon",     "Grade 10-A · SUN-S014"],
        ["STUDENT",      "siddharth.student",           "Demo@2026!",                "sunrise-academy", "Siddharth Rao",   "Grade 10-A · SUN-S015"],
        ["PARENT",       "ramesh.parent",               "Demo@2026!",                "sunrise-academy", "Ramesh Sharma",   "Parent of Aarav Sharma"],
        ["PARENT",       "sujata.parent",               "Demo@2026!",                "sunrise-academy", "Sujata Patel",    "Parent of Diya Patel"],
        ["PARENT",       "mukesh.parent",               "Demo@2026!",                "sunrise-academy", "Mukesh Singh",    "Parent of Kabir & Ananya"],
        ["PARENT",       "geeta.parent",                "Demo@2026!",                "sunrise-academy", "Geeta Roy",       "Parent of Rohan Kumar"],
        ["PARENT",       "naveen.parent",               "Demo@2026!",                "sunrise-academy", "Naveen Kumar",    "Parent of Priya & Aryan"],
        ["PARENT",       "anita.parent",                "Demo@2026!",                "sunrise-academy", "Anita Joshi",     "Parent of Sneha & Vikrant"],
        ["PARENT",       "rajesh.parent",               "Demo@2026!",                "sunrise-academy", "Rajesh Mehta",    "Parent of Riya Iyer"],
    ]
    write_table(ws, 4, headers, rows)
    set_col_widths(ws, [16, 28, 14, 18, 18, 34])

    # Role colour bands
    role_colors = {
        "SUPER_ADMIN":  "7C3AED",
        "SCHOOL_ADMIN": BLUE_DARK,
        "TEACHER":      "065F46",
        "STUDENT":      "1E40AF",
        "PARENT":       "92400E",
    }
    for row_idx in range(5, 5 + len(rows)):
        cell = ws.cell(row=row_idx, column=1)
        color = role_colors.get(cell.value, "000000")
        cell.font = Font(color=color, bold=True, size=10, name="Calibri")


# ════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_overview(wb)
    build_users(wb)
    build_teachers(wb)
    build_students(wb)
    build_parents(wb)
    build_classes(wb)
    build_subjects(wb)
    build_exams(wb)
    build_exam_results(wb)
    build_fees(wb)
    build_homework(wb)
    build_timetable(wb)
    build_attendance(wb)
    build_website(wb)
    build_admission_leads(wb)
    build_quick_login(wb)

    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "CloudCampus_DummyData.xlsx"
    )
    wb.save(out_path)
    print(f"Saved → {out_path}")


if __name__ == "__main__":
    main()

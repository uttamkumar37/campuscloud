#!/usr/bin/env python3
"""
CloudCampus — Dashboard Data Seed Script
Seeds one complete school (Sunrise Academy) with all data needed
to showcase the new STUDENT and TEACHER dashboards:
  classes, sections, subjects, teachers, students, timetable,
  attendance (30 days), fee assignments, exams, exam results,
  homework — all via the live API.
"""

import io
import json
import random
import sys
import time
import warnings
from datetime import date, timedelta

warnings.filterwarnings("ignore")

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing required packages …")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "openpyxl", "-q"])
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

BASE = "http://localhost:8080/api/v1"
TODAY = date.today()

# ── Credentials ────────────────────────────────────────────────────────────────
SA_USER = "superadmin"
SA_PASS = "SuperAdmin_Docker_2026!"

TENANT = {
    "tenantId":    "sunrise-academy",
    "schoolName":  "Sunrise Academy",
    "primaryColor": "#10b981",
    "logoUrl":     "https://images.unsplash.com/photo-1513258496099-48168024aec0?auto=format&fit=crop&w=200&q=80",
    "admin": {
        "fullName": "Priya Sharma",
        "username": "priya.sharma",
        "email":    "priya@sunrise.edu",
        "password": "Sunrise@2026!",
    },
    "code": "SUN",
}

SUBJECTS = [
    ("Mathematics",    "MATH"),
    ("Science",        "SCI"),
    ("English",        "ENG"),
    ("History",        "HIST"),
    ("Geography",      "GEO"),
    ("Physics",        "PHY"),
]

CLASSES = [
    ("Grade 1",  "SUN-G1"),
    ("Grade 3",  "SUN-G3"),
    ("Grade 5",  "SUN-G5"),
    ("Grade 7",  "SUN-G7"),
    ("Grade 10", "SUN-G10"),
]
SECTIONS = ["A", "B"]

STUDENT_FIRST = ["Aarav","Diya","Kabir","Ananya","Vivaan","Ishita","Aryan","Priya",
                 "Rohan","Sneha","Aditya","Meera","Siddharth","Pooja","Karan"]
STUDENT_LAST  = ["Sharma","Patel","Singh","Kumar","Verma","Reddy","Nair","Joshi",
                 "Mehta","Gupta","Iyer","Rao","Shah","Chowdhury","Malhotra"]

TEACHER_DATA = [
    ("Sunita",  "Aggarwal",  "Mathematics", "SCI"),
    ("Vikram",  "Desai",     "Science",     "SCI"),
    ("Lakshmi", "Krishnan",  "English",     "ENG"),
    ("Rajesh",  "Bose",      "History",     "HIST"),
    ("Anita",   "Pillai",    "Geography",   "GEO"),
    ("Deepak",  "Chauhan",   "Physics",     "PHY"),
]

# ── HTTP helper ────────────────────────────────────────────────────────────────
def api(method, path, token=None, tenant=None, json_body=None, files=None, quiet=False):
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if tenant:
        headers["X-Tenant-ID"] = tenant
    if json_body is not None and files is None:
        headers["Content-Type"] = "application/json"
    resp = requests.request(
        method, f"{BASE}{path}",
        headers=headers,
        json=json_body,
        files=files,
        timeout=30,
    )
    try:
        return resp.json()
    except Exception:
        return {"success": False, "message": resp.text[:300]}

def ok(r):
    return r.get("success") or r.get("data") is not None

def login(username, password, tenant=None):
    # tenant is a slug/schemaName — backend expects it in the JSON body as "tenantSlug"
    body = {"username": username, "password": password}
    if tenant:
        # strip the "school_" prefix if present to get the slug
        slug = tenant.removeprefix("school_") if tenant.startswith("school_") else tenant
        body["tenantSlug"] = slug
    r = requests.post(f"{BASE}/auth/login",
                      json=body,
                      headers={"Content-Type": "application/json"},
                      timeout=15)
    d = r.json()
    if d.get("success"):
        return d["data"]["accessToken"], d["data"].get("userId")
    return None, None

def get_me(token, tenant):
    r = api("GET", "/auth/me", token=token, tenant=tenant)
    return r.get("data", {}).get("userId")

def log(sym, msg):
    print(f"  {sym} {msg}")

# ── Excel builder ──────────────────────────────────────────────────────────────
def build_workbook(code, students_rows, teachers_rows, classes_rows, sections_rows):
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    center = Alignment(horizontal="center")

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for row in rows:
            ws.append(row)
        for col_idx in range(1, len(headers)+1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    add_sheet("STUDENTS", ["admission_no","first_name","last_name","dob","gender","email","phone"], students_rows)
    add_sheet("TEACHERS", ["employee_no","first_name","last_name","email","phone","hire_date"], teachers_rows)
    add_sheet("CLASSES",  ["class_name","class_code"], classes_rows)
    add_sheet("SECTIONS", ["section_name","class_code"], sections_rows)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── Grade calculator ───────────────────────────────────────────────────────────
def calc_grade(marks, max_marks):
    pct = (marks / max_marks) * 100
    if pct >= 90: return "A+"
    if pct >= 80: return "A"
    if pct >= 70: return "B+"
    if pct >= 60: return "B"
    if pct >= 50: return "C"
    return "D"

# ── Main ──────────────────────────────────────────────────────────────────────
def seed():
    print("\n" + "═"*65)
    print("  CloudCampus — Sunrise Academy Complete Data Seed")
    print("═"*65)

    # ── Step 1: Super-admin login ──────────────────────────────────
    print("\n[1] Super Admin login …")
    sa_token, _ = login(SA_USER, SA_PASS)
    if not sa_token:
        print("  ❌  Cannot login as superadmin. Is the backend running?")
        sys.exit(1)
    log("✅", "Logged in as Super Admin")

    # ── Step 2: Resolve plan ───────────────────────────────────────
    print("\n[2] Resolving subscription plan …")
    plans = api("GET", "/plans", token=sa_token).get("data", [])
    plan = next((p for p in plans if p["name"] in ("BASIC","FREE")), None)
    if not plan:
        log("❌", "No subscription plan found")
        sys.exit(1)
    plan_id = plan["id"]
    log("✅", f"Plan: {plan['name']} ({plan_id})")

    # ── Step 3: Create / resolve tenant ───────────────────────────
    print("\n[3] Setting up Sunrise Academy …")
    tid = TENANT["tenantId"]
    admin = TENANT["admin"]
    t_resp = api("POST", "/tenants", token=sa_token, json_body={
        "tenantId":    tid,
        "schoolName":  TENANT["schoolName"],
        "primaryColor": TENANT["primaryColor"],
        "logoUrl":     TENANT["logoUrl"],
        "schoolAdminFullName": admin["fullName"],
        "schoolAdminUsername": admin["username"],
        "schoolAdminEmail":    admin["email"],
        "schoolAdminPassword": admin["password"],
    })
    if ok(t_resp):
        schema = t_resp["data"]["schemaName"]
        log("✅", f"Tenant created → schema={schema}")
    else:
        msg = t_resp.get("message","")
        if "already" in msg.lower() or "duplicate" in msg.lower() or "exists" in msg.lower():
            tenants_list = api("GET", "/tenants", token=sa_token).get("data", [])
            existing = next((t for t in tenants_list if t["tenantId"] == tid), None)
            if not existing:
                log("❌", f"Cannot resolve tenant: {msg}")
                sys.exit(1)
            schema = existing["schemaName"]
            log("ℹ️ ", f"Tenant already exists → schema={schema}")
        else:
            log("❌", f"Tenant creation failed: {msg}")
            sys.exit(1)

    # ── Step 4: Subscribe ──────────────────────────────────────────
    sub_resp = api("POST", f"/tenants/{tid}/subscribe", token=sa_token,
                   json_body={"planId": plan_id, "durationDays": 365})
    if ok(sub_resp):
        log("✅", "Subscribed to plan")
    else:
        msg = sub_resp.get("message","")
        if "already" in msg.lower() or "active" in msg.lower():
            log("ℹ️ ", "Already subscribed")
        else:
            log("⚠️ ", f"Subscribe: {msg}")

    # ── Step 5: Create school admin user ──────────────────────────
    adm = TENANT["admin"]
    u_resp = api("POST", "/users", token=sa_token, tenant=schema, json_body={
        "fullName": adm["fullName"],
        "username": adm["username"],
        "email":    adm["email"],
        "password": adm["password"],
        "role":     "SCHOOL_ADMIN",
    })
    if ok(u_resp):
        log("✅", f"School admin created: {adm['username']}")
    elif "already" in str(u_resp.get("message","")).lower() or "duplicate" in str(u_resp.get("message","")).lower():
        log("ℹ️ ", f"Admin user exists: {adm['username']}")
    else:
        log("❌", f"Admin user: {u_resp.get('message')}")

    # ── Step 6: Login as school admin ─────────────────────────────
    print("\n[4] Logging in as school admin …")
    adm_token, _ = login(adm["username"], adm["password"], tenant=schema)
    if not adm_token:
        log("❌", "School admin login failed!")
        sys.exit(1)
    adm_user_id = get_me(adm_token, schema)
    log("✅", f"Logged in (userId={adm_user_id})")

    # ── Step 7: Bulk upload (students, teachers, classes, sections) ─
    print("\n[5] Bulk uploading base data …")
    code = TENANT["code"]
    students_rows, teachers_rows, classes_rows, sections_rows = [], [], [], []
    student_creds, teacher_creds = [], []

    for c_name, c_code in CLASSES:
        classes_rows.append([c_name, c_code])
        for sec in SECTIONS:
            sections_rows.append([sec, c_code])

    for i, (fn, ln) in enumerate(zip(STUDENT_FIRST, STUDENT_LAST), 1):
        adm_no = f"{code}-S{i:03d}"
        students_rows.append([
            adm_no, fn, ln,
            f"{2010-(i%5)}-{(i%12)+1:02d}-{(i%28)+1:02d}",
            "MALE" if i%2 else "FEMALE",
            f"{fn.lower()}.{ln.lower()}{i}@sunrise.edu",
            f"98765{i:05d}",
        ])
        student_creds.append({
            "admNo": adm_no, "name": f"{fn} {ln}",
            "username": f"{fn.lower()}.{ln.lower()}{i}",
            "password": f"{fn}@Student{i}26!",
            "email": f"{fn.lower()}.{ln.lower()}{i}@sunrise.edu",
        })

    for i, (fn, ln, subj, subj_code) in enumerate(TEACHER_DATA, 1):
        emp_no = f"{code}-T{i:02d}"
        teachers_rows.append([
            emp_no, fn, ln,
            f"{fn.lower()}.{ln.lower()}@sunrise.edu",
            f"91234{i:05d}",
            f"202{i%4}-0{i}-15",
        ])
        teacher_creds.append({
            "empNo": emp_no, "name": f"{fn} {ln}",
            "subject": subj, "subjectCode": subj_code,
            "username": f"{fn.lower()}.{ln.lower()}",
            "password": f"{fn}@Teacher{i}26!",
            "email": f"{fn.lower()}.{ln.lower()}@sunrise.edu",
        })

    xlsx = build_workbook(code, students_rows, teachers_rows, classes_rows, sections_rows)
    up = api("POST", "/bulk/upload", token=adm_token, tenant=schema,
             files={"file": ("bulk.xlsx", xlsx,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    if ok(up):
        d = up.get("data",{})
        log("✅", f"Bulk upload: {d.get('successCount')}/{d.get('totalRows')} OK, {d.get('failedCount')} failed")
    else:
        log("⚠️ ", f"Bulk upload: {up.get('message','unknown error')}")

    # ── Step 8: Create teacher + student user accounts ─────────────
    print("\n[6] Creating user accounts …")
    for tc in teacher_creds:
        u = api("POST", "/users", token=adm_token, tenant=schema, json_body={
            "fullName": tc["name"], "username": tc["username"],
            "email": tc["email"], "password": tc["password"], "role": "TEACHER",
        })
        if ok(u):
            tc["userId"] = u["data"]["id"]
            log("✅", f"Teacher: {tc['username']}")
        elif "already" in str(u.get("message","")).lower() or "duplicate" in str(u.get("message","")).lower():
            log("ℹ️ ", f"Teacher exists: {tc['username']}")
        else:
            log("⚠️ ", f"Teacher {tc['username']}: {u.get('message')}")

    for sc in student_creds[:5]:
        u = api("POST", "/users", token=adm_token, tenant=schema, json_body={
            "fullName": sc["name"], "username": sc["username"],
            "email": sc["email"], "password": sc["password"], "role": "STUDENT",
        })
        if ok(u):
            sc["userId"] = u["data"]["id"]
            log("✅", f"Student: {sc['username']}")
        elif "already" in str(u.get("message","")).lower() or "duplicate" in str(u.get("message","")).lower():
            log("ℹ️ ", f"Student exists: {sc['username']}")
        else:
            log("⚠️ ", f"Student {sc['username']}: {u.get('message')}")

    # ── Step 9: Fetch all IDs ──────────────────────────────────────
    print("\n[7] Fetching entity IDs …")

    classes_resp = api("GET", "/academics/classes", token=adm_token, tenant=schema)
    all_classes = classes_resp.get("data", [])
    class_by_code = {c["code"]: c for c in all_classes}
    log("✅", f"Classes: {len(all_classes)}")

    sections_resp = api("GET", "/academics/sections", token=adm_token, tenant=schema)
    all_sections = sections_resp.get("data", [])
    log("✅", f"Sections: {len(all_sections)}")

    students_resp = api("GET", "/students?page=0&size=20", token=adm_token, tenant=schema)
    all_students = students_resp.get("data", {}).get("content", [])
    log("✅", f"Students: {len(all_students)}")

    teachers_resp = api("GET", "/teachers?page=0&size=10", token=adm_token, tenant=schema)
    all_teachers = teachers_resp.get("data", {}).get("content", [])
    log("✅", f"Teachers: {len(all_teachers)}")

    if not all_classes or not all_sections or not all_students or not all_teachers:
        log("❌", "Missing base data. Bulk upload may have failed. Check manually.")
        sys.exit(1)

    # Re-fetch teacher userIds by logging in
    for tc in teacher_creds:
        if not tc.get("userId"):
            tok, uid = login(tc["username"], tc["password"], tenant=schema)
            if tok and uid:
                tc["userId"] = uid
            else:
                # Try to find by employeeNo
                match = next((t for t in all_teachers if t.get("employeeNo") == tc["empNo"]), None)
                if match:
                    tc["teacherEntityId"] = match["id"]

    # Match teacher entity IDs
    teacher_by_empno = {t["employeeNo"]: t for t in all_teachers}
    for tc in teacher_creds:
        match = teacher_by_empno.get(tc["empNo"])
        if match:
            tc["teacherEntityId"] = match["id"]

    # Match student entity IDs (to first 5 with user accounts)
    student_by_admno = {s["admissionNo"]: s for s in all_students}
    for sc in student_creds:
        match = student_by_admno.get(sc["admNo"])
        if match:
            sc["studentEntityId"] = match["id"]

    # Sections grouped by class (section has className field, not nested schoolClass)
    sections_by_class_name = {}
    for s in all_sections:
        cn = s.get("className") or s.get("schoolClass", {}).get("name", "")
        sections_by_class_name.setdefault(cn, []).append(s)

    # ── Step 10: Create subjects ───────────────────────────────────
    print("\n[8] Creating subjects …")
    subject_id_map = {}  # code -> id
    for name, code_str in SUBJECTS:
        r = api("POST", "/academics/subjects", token=adm_token, tenant=schema,
                json_body={"name": name, "code": code_str})
        if ok(r):
            subject_id_map[code_str] = r["data"]["id"]
            log("✅", f"Subject: {name}")
        elif "already" in str(r.get("message","")).lower() or "duplicate" in str(r.get("message","")).lower() or "exists" in str(r.get("message","")).lower():
            log("ℹ️ ", f"Subject exists: {name}")
        else:
            log("⚠️ ", f"Subject {name}: {r.get('message')}")

    # Fetch subjects to get IDs for existing ones
    subs_resp = api("GET", "/academics/subjects", token=adm_token, tenant=schema)
    for s in subs_resp.get("data", []):
        subject_id_map[s["code"]] = s["id"]
    log("✅", f"Subject IDs resolved: {len(subject_id_map)}")

    # ── Step 11: Create timetable slots ───────────────────────────
    print("\n[9] Creating timetable slots …")
    # Assign one teacher per subject across Grade 5A and Grade 7A as primary classes
    # Each teacher gets 2 periods per day in their class
    # dayOfWeek: 1=Mon, 2=Tue, 3=Wed, 4=Thu, 5=Fri

    # Find Grade 5 and Grade 7 class IDs and Section A
    g5 = next((c for c in all_classes if "5" in c.get("name","")), None)
    g7 = next((c for c in all_classes if "7" in c.get("name","")), None)
    g10 = next((c for c in all_classes if "10" in c.get("name","")), None)

    timetable_slots_created = 0

    for cls in [g5, g7, g10]:
        if not cls:
            continue
        cls_id = cls["id"]
        cls_sections = sections_by_class_name.get(cls["name"], [])
        if not cls_sections:
            continue
        sec_a = next((s for s in cls_sections if s["name"] == "A"), cls_sections[0])
        sec_id = sec_a["id"]

        # 6 teachers × 5 subjects (6 subjects, 5 days = 1 period per teacher per day)
        periods = [
            ("08:00", "08:45"), ("08:45", "09:30"), ("09:30", "10:15"),
            ("10:30", "11:15"), ("11:15", "12:00"), ("12:00", "12:45"),
        ]
        for day in range(1, 6):  # Mon-Fri
            for idx, (tc, (fn, ln, subj_name, subj_code)) in enumerate(zip(teacher_creds, TEACHER_DATA)):
                sub_id = subject_id_map.get(subj_code)
                teacher_id = tc.get("teacherEntityId")
                if not sub_id:
                    continue
                start_t, end_t = periods[idx % len(periods)]
                slot_req = {
                    "classId":   cls_id,
                    "sectionId": sec_id,
                    "subjectId": sub_id,
                    "dayOfWeek": day,
                    "startTime": f"{start_t}:00",
                    "endTime":   f"{end_t}:00",
                    "label":     f"Period {idx+1}",
                }
                if teacher_id:
                    slot_req["teacherId"] = teacher_id
                r = api("POST", "/timetable/slots", token=adm_token, tenant=schema, json_body=slot_req)
                if ok(r):
                    timetable_slots_created += 1

    log("✅", f"Timetable slots created: {timetable_slots_created}")

    # ── Step 12: Mark attendance for last 30 days ──────────────────
    print("\n[10] Marking attendance (last 30 days) …")
    if not g5 or not g7:
        log("⚠️ ", "Grade 5 / 7 not found, skipping attendance")
    else:
        g5_sections = sections_by_class_name.get(g5["name"], [])
        g5_sec_a = next((s for s in g5_sections if s["name"] == "A"), g5_sections[0] if g5_sections else None)

        att_total = 0
        if g5_sec_a:
            for student in all_students[:8]:  # First 8 students
                sid = student["id"]
                for offset in range(30):
                    att_date = TODAY - timedelta(days=offset)
                    if att_date.weekday() >= 5:  # skip weekends
                        continue
                    # Realistic attendance pattern
                    rng = random.Random(sid + str(att_date))
                    roll = rng.random()
                    status = "PRESENT" if roll > 0.12 else ("ABSENT" if roll > 0.04 else ("LATE" if roll > 0.01 else "EXCUSED"))
                    r = api("POST", "/attendances", token=adm_token, tenant=schema, json_body={
                        "studentId":      sid,
                        "classId":        g5["id"],
                        "sectionId":      g5_sec_a["id"],
                        "attendanceDate": str(att_date),
                        "status":         status,
                        "markedByUserId": adm_user_id,
                    })
                    if ok(r):
                        att_total += 1
        log("✅", f"Attendance records created: {att_total}")

    # ── Step 13: Create fee assignments ───────────────────────────
    print("\n[11] Creating fee assignments …")
    fee_total = 0
    fee_data = [
        ("Annual Tuition Fee", 15000, 30),
        ("Exam Fee - Term 1",   1500, 15),
        ("Lab Fee",             2000, 20),
        ("Sports Fee",           500, 45),
        ("Library Fee",          300, 60),
    ]
    fee_ids = []
    for student in all_students[:8]:
        sid = student["id"]
        for title, amount, due_offset in fee_data:
            due = TODAY + timedelta(days=due_offset)
            r = api("POST", "/fees/assignments", token=adm_token, tenant=schema, json_body={
                "studentId": sid,
                "feeTitle":  title,
                "amount":    amount,
                "dueDate":   str(due),
            })
            if ok(r):
                fee_ids.append(r["data"]["id"])
                fee_total += 1
    log("✅", f"Fee assignments created: {fee_total}")

    # ── Step 14: Record some fee payments ─────────────────────────
    print("\n[12] Recording fee payments …")
    pay_total = 0
    for fee_id in fee_ids[:10]:  # Pay first 10 assignments
        r = api("POST", "/fees/payments", token=adm_token, tenant=schema, json_body={
            "feeAssignmentId": fee_id,
            "amountPaid":      1500,
            "paymentDate":     str(TODAY - timedelta(days=random.randint(1, 20))),
            "paymentMethod":   random.choice(["CASH", "UPI", "BANK_TRANSFER"]),
            "referenceNo":     f"REF{random.randint(100000,999999)}",
            "receivedByUserId": adm_user_id,
        })
        if ok(r):
            pay_total += 1
    log("✅", f"Payments recorded: {pay_total}")

    # ── Step 15: Create exams ──────────────────────────────────────
    print("\n[13] Creating exams …")
    exam_ids = []
    exam_data = [
        ("Mathematics Mid-Term",   TODAY - timedelta(days=20), "MATH", 100),
        ("Science Mid-Term",       TODAY - timedelta(days=18), "SCI",  80),
        ("English Mid-Term",       TODAY - timedelta(days=15), "ENG",  100),
        ("History Unit Test",      TODAY - timedelta(days=10), "HIST", 50),
        ("Geography Unit Test",    TODAY - timedelta(days=8),  "GEO",  50),
        ("Physics Practical",      TODAY - timedelta(days=5),  "PHY",  40),
    ]

    if g5 and g5_sections:
        g5_sec_a_id = next((s["id"] for s in g5_sections if s["name"] == "A"), g5_sections[0]["id"])
        for title, exam_date, subj_code, max_marks in exam_data:
            sub_id = subject_id_map.get(subj_code)
            if not sub_id:
                continue
            r = api("POST", "/exams", token=adm_token, tenant=schema, json_body={
                "title":    title,
                "examDate": str(exam_date),
                "classId":  g5["id"],
                "sectionId": g5_sec_a_id,
                "subjectId": sub_id,
                "maxMarks": max_marks,
            })
            if ok(r):
                exam_ids.append({"id": r["data"]["id"], "maxMarks": max_marks, "title": title})
                log("✅", f"Exam: {title}")
            else:
                log("⚠️ ", f"Exam {title}: {r.get('message')}")

    # ── Step 16: Create exam results ──────────────────────────────
    print("\n[14] Creating exam results …")
    result_total = 0
    rng = random.Random(42)
    for exam in exam_ids:
        for student in all_students[:8]:
            sid = student["id"]
            max_m = exam["maxMarks"]
            marks = round(rng.uniform(max_m * 0.45, max_m * 0.98), 1)
            grade = calc_grade(marks, max_m)
            r = api("POST", "/exams/results", token=adm_token, tenant=schema, json_body={
                "examId":       exam["id"],
                "studentId":    sid,
                "marksObtained": marks,
                "grade":        grade,
                "remarks":      f"{'Excellent' if grade in ('A+','A') else 'Good' if grade in ('B+','B') else 'Needs improvement'}",
                "published":    True,
            })
            if ok(r):
                result_total += 1
    log("✅", f"Exam results created: {result_total}")

    # ── Step 17: Create homework assignments ──────────────────────
    print("\n[15] Creating homework assignments …")
    hw_total = 0
    hw_data = [
        ("Chapter 5 Practice Problems", "Solve all exercises from Chapter 5 of the textbook.", "MATH", 3),
        ("Lab Report — Photosynthesis", "Write a 2-page lab report on the photosynthesis experiment.", "SCI", 5),
        ("Essay: My Favourite Season", "Write a 300-word descriptive essay.", "ENG", 4),
        ("Timeline: Freedom Struggle",  "Create a visual timeline of India's freedom movement.", "HIST", 7),
        ("Map Work: South Asia",        "Label all countries and capitals of South Asia on a blank map.", "GEO", 3),
        ("Numericals: Newton's Laws",   "Solve the numericals on page 78-79 of Physics textbook.", "PHY", 2),
    ]

    if g5:
        for idx, (title, instructions, subj_code, due_days) in enumerate(hw_data):
            due = TODAY + timedelta(days=due_days)
            r = api("POST", "/homework", token=adm_token, tenant=schema, json_body={
                "title":        title,
                "instructions": instructions,
                "classId":      g5["id"],
                "sectionId":    g5_sec_a_id if g5_sections else None,
                "dueDate":      str(due),
            })
            if ok(r):
                hw_total += 1
                log("✅", f"Homework: {title}")
            else:
                log("⚠️ ", f"Homework {title}: {r.get('message')}")

    # Also create homework as teacher (for teacher dashboard)
    for tc in teacher_creds[:3]:
        tok, _ = login(tc["username"], tc["password"], tenant=schema)
        if not tok or not g5:
            continue
        title = f"{tc['subject']} — Practice Set {rng.randint(1,10)}"
        due = TODAY + timedelta(days=rng.randint(2, 8))
        r = api("POST", "/homework", token=tok, tenant=schema, json_body={
            "title":        title,
            "instructions": f"Complete all exercises assigned for {tc['subject']}.",
            "classId":      g5["id"],
            "sectionId":    g5_sec_a_id if g5_sections else None,
            "dueDate":      str(due),
        })
        if ok(r):
            hw_total += 1

    log("✅", f"Homework assignments created: {hw_total}")

    # ── Summary ────────────────────────────────────────────────────
    print("\n" + "═"*65)
    print("  SEED COMPLETE — Login Credentials")
    print("═"*65)

    print(f"""
  🏫  School: Sunrise Academy
  🌐  Login URL: http://localhost:5173/login
  📋  X-Tenant-ID: {schema}

  ─── SCHOOL ADMIN ─────────────────────────────────────────
  username : priya.sharma
  password : Sunrise@2026!

  ─── TEACHERS ─────────────────────────────────────────────""")
    for tc in teacher_creds:
        print(f"  {tc['username']:<30}  {tc['password']:<25}  ({tc['subject']})")

    print(f"""
  ─── STUDENTS (with dashboard login) ──────────────────────""")
    for sc in student_creds[:5]:
        print(f"  {sc['username']:<30}  {sc['password']:<25}  adm:{sc['admNo']}")

    print(f"""
  ─── DASHBOARD URLS ───────────────────────────────────────
  Student dashboard : http://localhost:5173/student/dashboard
  Teacher dashboard : http://localhost:5173/teacher/dashboard
  Admin dashboard   : http://localhost:5173/dashboard

  ─── SUPER ADMIN ──────────────────────────────────────────
  URL      : http://localhost:5173/super-admin/login
  username : superadmin
  password : SuperAdmin_Docker_2026!
""")
    print("═"*65 + "\n")

if __name__ == "__main__":
    seed()
